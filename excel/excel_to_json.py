import json
import os
import re
from typing import Union, List, Dict, Optional

from openpyxl import load_workbook
import logging as lg

logger = lg.getLogger("OrderExport")
logger.setLevel(lg.INFO)
formatter = lg.Formatter('%(asctime)s - %(levelname)s - %(message)s')

def setup_logger(log_dir: str) -> None:
    log_path = os.path.join(log_dir, "export_orders.log")
    file_handler = lg.FileHandler(log_path, encoding='utf-8')
    file_handler.setFormatter(formatter)
    logger.addHandler(file_handler)


def extract_type_details(line_item: str) -> Optional[str]:
    if not line_item or not isinstance(line_item, str):
        return None

    line_item = line_item.strip()

    # תבנית 1: מדבקות שם - חברים - סט מדבקות 52+90
    pattern1 = r"^מדבקות שם(?:\s*-\s*.+)*?\s*-\s*(?P<design>[^-]+?)\s*-\s*(?:סט\s+מדבקות\s*)?(?P<quantity>\d+(?:\+\d+)?)"

    # תבנית 2: מדבקות שם קטנות במיוחד - ברקע חד קרן ללא איורים
    pattern2 = r"ברקע\s+(?P<design>.+?)\s+ללא\s+איורים"

    # תבנית 3: שקופות עם הדפסה ללא איורים
    pattern3 = r"-\s*(?P<design>[^-]+?)\s+ללא\s+איורים"

    match1 = re.search(pattern1, line_item)
    if match1:
        design = match1.group("design").strip()
        quantity = match1.group("quantity").strip()
        return f"{design}_{quantity}"

    match2 = re.search(pattern2, line_item)
    if match2:
        design = match2.group("design").strip()
        quantity_match = re.search(r"(\d+(?:\+\d+)?)", line_item)
        quantity = quantity_match.group(1) if quantity_match else "unknown"
        return f"{design}_{quantity}"

    match3 = re.search(pattern3, line_item)
    if match3:
        design = match3.group("design").strip()
        quantity_match = re.search(r"(\d+(?:\+\d+)?)", line_item)
        quantity = quantity_match.group(1) if quantity_match else "unknown"
        return f"{design}_{quantity}"

    return None


def extract_name_details(line_item: str) -> Optional[str]:

    if not line_item:
        return None

    match = re.search(
        r"שם היל[ד|ד/ה|דה|דה/ה]* שיודפס על גבי המדבקות:\s*(.+?)\s*(?:\n|$)",
        line_item,
        re.IGNORECASE
    )
    if match:
        name = match.group(1).strip()
        return name if name else None

    return " "


def split_item_variants(item: Dict) -> List[Dict]:

    item_name = item.get("itemName", "")
    if isinstance(item_name, str) and "_52+90" in item_name:
        base_name = item_name.replace("_52+90", "")
        item_52 = item.copy()
        item_52["itemName"] = f"{base_name}_52"
        item_90 = item.copy()
        item_90["itemName"] = f"{base_name}_90"
        return [item_52, item_90]
    return [item]



def excel_to_filtered_json(excel_file_path: str) -> list[list[dict]] | None:

    try:
        wb=load_workbook(excel_file_path)
        ws=wb.active

        rows = list(ws.iter_rows(values_only=True))
        if not rows or len(rows) < 2:
            logger.warning("Excel file is empty or missing data.")
            return None

        headers = rows[0]

        field_mapping = {
            "Business Partner Reference Number": "referenceNumber",
            "Item Name": "itemName",
            "Quantity": "quantity",
            "Line Comments":"name"
        }

        header_indexes = {field_mapping[key]: headers.index(key) for key in field_mapping if key in headers}

        filtered_data = []
        for row in rows[1:]:
            item = {
                new_key: row[idx] for new_key, idx in header_indexes.items() if idx < len(row)
            }

            item_line_raw_type = item.get("itemName")
            item_details = extract_type_details(item_line_raw_type)
            if item_details:
                item["itemName"] = item_details

            item_line_raw_name = item.get("name")
            item_name=extract_name_details(item_line_raw_name)
            if item_name:
                item["name"]=item_name
            else:
                item["name"]=" "

            final_item= split_item_variants(item)
            filtered_data.append(final_item)

        base_name = os.path.splitext(excel_file_path)[0]
        output_json_path = base_name + ".json"
        if output_json_path:
            with open(output_json_path, 'w', encoding='utf-8') as f:
                json.dump(filtered_data, f, ensure_ascii=False, indent=2)
            logger.info(f"Exported filtered JSON to {output_json_path}")

        return filtered_data

    except Exception as e:
        logger.exception(f"Failed to convert Excel to filtered JSON: {e}")
        return None



if __name__ == '__main__':
    excel_to_filtered_json(r"C:\Users\USER\Downloads\Expected_Goods_Shipment_25_06_2025.xlsx")