import json
import os
import unittest
from openpyxl  import Workbook,load_workbook
from excel.excel_writer import parse_json_orders, split_orders, update_excel_row_by_dbid


class MyTestCase(unittest.TestCase):
    def setUp(self):
        self.test_file = r"C:\Users\USER\Downloads\Expected_Goods_Shipment_25_06_2025 (2).xlsx"
        wb = Workbook()
        ws = wb.active
        ws.append(["dbId", "graphicStatus", "orderStatus"])
        ws.append([123, "OldGraphic", "OldOrder"])
        wb.save(self.test_file)

    def tearDown(self):
        if os.path.exists(self.test_file):
            os.remove(self.test_file)

    def test_update_excel_row_success(self):
        update_excel_row_by_dbid(
            file_path=self.test_file,
            dbid=123,
            updates={"graphicStatus": "NewGraphic", "orderStatus": "NewOrder"}
        )
        wb = load_workbook(self.test_file)
        ws = wb.active
        self.assertEqual(ws["B2"].value, "NewGraphic")
        self.assertEqual(ws["C2"].value, "NewOrder")

    def test_update_excel_row_not_found(self):
        update_excel_row_by_dbid(
            file_path=self.test_file,
            dbid=999,
            updates={"graphicStatus": "NewGraphic"}
        )
        wb = load_workbook(self.test_file)
        ws = wb.active
        self.assertEqual(ws["B2"].value, "OldGraphic")
    def test_parse_json_orders_with_invalid_json(self):
        bad_json = '{"not": "a list"}'
        result = parse_json_orders(bad_json)
        self.assertIsNone(result)

    def test_parse_json_orders_with_valid_list(self):
        good_data = [{"orderId": 1}, {"orderId": 2}]
        result = parse_json_orders(json.dumps(good_data))
        self.assertEqual(result, good_data)

    def test_split_orders_exact_chunks(self):
        orders = [{"id": i} for i in range(10)]
        chunks = split_orders(orders, 5)
        self.assertEqual(len(chunks), 2)
        self.assertEqual(len(chunks[0]), 5)

    def test_split_orders_uneven_chunks(self):
        orders = [{"id": i} for i in range(13)]
        chunks = split_orders(orders, 5)
        self.assertEqual(len(chunks), 3)
        self.assertEqual(len(chunks[-1]), 3)

    def test_split_orders_empty(self):
        self.assertEqual(split_orders([], 5), [])


if __name__ == '__main__':
    unittest.main()
