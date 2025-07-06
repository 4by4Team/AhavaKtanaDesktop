import os
import json
from openpyxl import Workbook
from datetime import datetime
from openpyxl import load_workbook
from typing import Union, List, Dict, Optional
import logging as lg
import re

logger = lg.getLogger("OrderExport")
logger.setLevel(lg.INFO)
formatter = lg.Formatter('%(asctime)s - %(levelname)s - %(message)s')

def setup_logger(log_dir: str) -> None:
    log_path = os.path.join(log_dir, "export_orders.log")
    file_handler = lg.FileHandler(log_path, encoding='utf-8')
    file_handler.setFormatter(formatter)
    logger.addHandler(file_handler)



def parse_json_orders(orders: Union[str, List[Dict]]) -> Union[List[Dict], None]:
    if isinstance(orders, str):
        try:
            orders = json.loads(orders)
        except json.JSONDecodeError:
            logger.error("Invalid JSON format.")
            return None

    if not isinstance(orders, list) or not all(isinstance(o, dict) for o in orders):
        logger.error("Invalid structure – must be a list of dictionaries.")
        return None

    if not orders:
        logger.warning("No orders provided.")
        return None

    return orders


def create_output_folder(base_dir: str) -> str:
    today = datetime.today().strftime("%Y-%m-%d")
    full_path = os.path.join(base_dir, today)
    os.makedirs(full_path, exist_ok=True)
    setup_logger(full_path)
    logger.info(f"Output directory created: {full_path}")
    return full_path


def split_orders(orders: List[Dict], chunk_size: int) -> List[List[Dict]]:
    return [orders[i:i + chunk_size] for i in range(0, len(orders), chunk_size)]


def write_excel_file(orders_chunk: List[Dict], output_path: str, file_index: int) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Orders"

    headers = list(orders_chunk[0].keys())
    ws.append(headers)

    for order in orders_chunk:
        ws.append([order.get(k, "") for k in headers])

    today = datetime.today().strftime("%Y-%m-%d")
    filename = f"orders_{today}_{file_index}.xlsx"
    full_path = os.path.join(output_path, filename)
    wb.save(full_path)
    logger.info(f"Saved file: {full_path}")


def save_orders_to_excel( orders: Union[str, List[Dict]],
                          output_dir: str = "data",
                          max_per_file: int = 24) -> bool:
    try:
        parsed_orders = parse_json_orders(orders)
        if parsed_orders is None:
            return False

        folder_path = create_output_folder(output_dir)
        chunks = split_orders(parsed_orders, max_per_file)

        for i, chunk in enumerate(chunks, start=1):
            write_excel_file(chunk, folder_path, i)
        logger.info(f"Successfully saved {len(chunks)} Excel file(s).")
        return True

    except Exception as e:
        logger.exception(f"Unexpected error occurred: {e}")
        return False




def load_excel_headers(ws) -> Dict[str, int]:
    return {cell.value: idx for idx, cell in enumerate(ws[1], start=1)}


def find_row_by_dbid(ws, dbid: Union[int, str], dbid_col_idx: int):
    for row in ws.iter_rows(min_row=2):
        if str(row[dbid_col_idx - 1].value) == str(dbid):
            return row
    return None


def update_excel_row_by_dbid(
    file_path: str,
    dbid: Union[int, str],
    updates: Dict[str, Union[str, int, float]],
    save_as: Optional[str] = None
) -> None:

    try:
        wb = load_workbook(file_path)
        ws = wb.active

        headers = load_excel_headers(ws)

        if "dbId" not in headers:
            raise ValueError("Missing required column: 'dbId'")

        dbid_col_idx = headers["dbId"]
        target_row = find_row_by_dbid(ws, dbid, dbid_col_idx)

        if target_row is None:
            msg = f"dbId {dbid} not found in file."
            logger.warning(msg)
            print(msg)
            return

        for column, new_value in updates.items():
            if column not in headers:
                raise ValueError(f"Column '{column}' not found in file.")
            col_idx = headers[column]
            target_row[col_idx - 1].value = new_value
            logger.info(f"dbId={dbid}: '{column}' updated to '{new_value}'")

        output_path = save_as or file_path
        wb.save(output_path)
        logger.info(f"File saved: {output_path}")
        print(f"Row with dbId={dbid} updated successfully.")

    except FileNotFoundError:
        msg = f"File not found: {file_path}"
        logger.error(msg)
        print(f"Error: {msg}")

    except Exception as e:
        logger.exception("Unexpected error occurred while updating row.")
        print(f"Unexpected error: {e}")


def update_excel_column_by_dbid(
    file_path: str,
    dbids_to_update: List[Union[str, int]],
    column_name: str,
    new_value: str,
    save_as: str = None
) -> None:

    try:
        wb = load_workbook(file_path)
        ws = wb.active

        headers = {cell.value: idx for idx, cell in enumerate(ws[1], start=1)}
        if column_name not in headers or "dbId" not in headers:
            error_msg = f"Missing columns: dbId or {column_name}"
            logger.error(error_msg)
            raise ValueError(error_msg)

        col_idx = headers[column_name]
        dbid_idx = headers["dbId"]

        updated_rows = 0
        for row in ws.iter_rows(min_row=2):
            dbid_cell = row[dbid_idx - 1].value
            if str(dbid_cell) in map(str, dbids_to_update):
                row[col_idx - 1].value = new_value
                updated_rows += 1

        if save_as:
            wb.save(save_as)
            logger.info(f"File saved as {save_as}")
        else:
            wb.save(file_path)
            logger.info(f"File overwritten: {file_path}")

        logger.info(f"{updated_rows} rows updated successfully (column: {column_name})")
        print(f"{updated_rows} rows updated successfully.")

    except FileNotFoundError:
        logger.error(f"File not found: {file_path}")
        print(f"Error: File not found: {file_path}")

    except Exception as e:
        logger.exception("An unexpected error occurred while updating Excel.")
        print(f"Unexpected error: {e}")






