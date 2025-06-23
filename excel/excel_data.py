from typing import List, Dict
from openpyxl import load_workbook


def excel_to_dict(file_path: str) -> List[Dict[str, str]]:

    wb = load_workbook(file_path)
    ws = wb.active

    headers = [cell.value for cell in ws[1]]
    data = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        row_dict = {headers[i]: row[i] for i in range(len(headers))}
        data.append(row_dict)

    return data