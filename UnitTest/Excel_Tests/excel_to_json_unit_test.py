import os
import unittest
from openpyxl import Workbook

from tempfile import NamedTemporaryFile

from excel.excel_to_json import excel_to_filtered_json, split_item_variants, extract_name_details, extract_type_details


class MyTestCase(unittest.TestCase):
    def test_extract_type_details_valid_patterns(self):
        self.assertEqual(
            extract_type_details("מדבקות שם - חברים - סט מדבקות 52+90"),
            "חברים_52+90"
        )
        self.assertEqual(
            extract_type_details("מדבקות שם קטנות במיוחד - ברקע חד קרן ללא איורים"),
            "חד קרן_unknown"
        )
        self.assertEqual(
            extract_type_details("שקופות - קשת בענן ללא איורים 90"),
            "קשת בענן_90"
        )

    def test_extract_type_details_invalid_and_edge_cases(self):
        self.assertIsNone(extract_type_details(None))
        self.assertIsNone(extract_type_details(""))
        self.assertIsNone(extract_type_details("לא תואם שום תבנית"))
        self.assertIsNone(extract_type_details(12345))

    def test_extract_name_details_valid_and_edge_cases(self):
        self.assertEqual(
            extract_name_details("שם הילד שיודפס על גבי המדבקות: יוסי"),
            "יוסי"
        )
        self.assertEqual(
            extract_name_details("שם הילדה שיודפס על גבי המדבקות: שירה\nהערה"),
            "שירה"
        )
        self.assertEqual(extract_name_details("שם הילד שיודפס על גבי המדבקות:   "), " ")
        self.assertEqual(extract_name_details(""), " ")
        self.assertEqual(extract_name_details(None), None)

    def test_split_item_variants_with_52_plus_90(self):
        item = {"itemName": "חד קרן_52+90", "other": "value"}
        result = split_item_variants(item)
        self.assertEqual(len(result), 2)
        self.assertEqual(result[0]["itemName"], "חד קרן_52")
        self.assertEqual(result[1]["itemName"], "חד קרן_90")

    def test_split_item_variants_without_52_plus_90(self):
        item = {"itemName": "חד קרן_90", "other": "value"}
        result = split_item_variants(item)
        self.assertEqual(len(result), 1)
        self.assertEqual(result[0]["itemName"], "חד קרן_90")

    def test_excel_to_filtered_json_basic_flow(self):
        wb = Workbook()
        ws = wb.active
        ws.append([
            "Business Partner Reference Number",
            "Item Name",
            "Quantity",
            "Line Comments"
        ])
        ws.append([
            "ABC123",
            "מדבקות שם - חד קרן - סט מדבקות 52+90",
            1,
            "שם הילד שיודפס על גבי המדבקות: יוסי"
        ])

        with NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
            temp_path = tmp.name
            wb.save(temp_path)

        result = excel_to_filtered_json(temp_path)

        self.assertIsInstance(result, list)
        self.assertEqual(len(result), 1)
        self.assertEqual(len(result[0]), 2)  # 52 + 90
        self.assertIn("itemName", result[0][0])
        self.assertEqual(result[0][0]["name"], "יוסי")

        json_path = temp_path.replace(".xlsx", ".json")
        self.assertTrue(os.path.exists(json_path))

        os.remove(temp_path)
        os.remove(json_path)

    def test_excel_to_filtered_json_empty_file(self):
        wb = Workbook()
        ws = wb.active
        with NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
            path = tmp.name
            wb.save(path)

        result = excel_to_filtered_json(path)
        self.assertIsNone(result)
        os.remove(path)

    def test_excel_to_filtered_json_missing_headers(self):
        wb = Workbook()
        ws = wb.active
        ws.append(["Invalid", "Headers"])
        with NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
            path = tmp.name
            wb.save(path)

        result = excel_to_filtered_json(path)
        self.assertIsNone(result)
        os.remove(path)




if __name__ == '__main__':
    unittest.main()
